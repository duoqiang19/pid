#!/usr/bin/env python
# -*- coding: utf-8 -*-
#
#  pid_handle_drawing_database.py
#
#  Copyright 2020 Wang D.Q <Wang D.Q@DESKTOP-LTKS88H>
#
#  This program is free software; you can redistribute it and/or modify
#  it under the terms of the GNU General Public License as published by
#  the Free Software Foundation; either version 2 of the License, or
#  (at your option) any later version.
#
#  This program is distributed in the hope that it will be useful,
#  but WITHOUT ANY WARRANTY; without even the implied warranty of
#  MERCHANTABILITY or FITNESS FOR A PARTICULAR PURPOSE.  See the
#  GNU General Public License for more details.
#
#  You should have received a copy of the GNU General Public License
#  along with this program; if not, write to the Free Software
#  Foundation, Inc., 51 Franklin Street, Fifth Floor, Boston,
#  MA 02110-1301, USA.
#
#
import xlrd
from xlutils.copy import copy  # 导入copy模块
import openpyxl

class Handle_drawing_info():

    def __init__(self, address):
        self.address = address
        self.book = openpyxl.load_workbook(self.address)
        self.sheet = self.book['图纸']
        self.rows_no = self.sheet.max_row
        self.cols_no = self.sheet.max_column
        #self.book = xlrd.open_workbook(self.address)
        #self.sheet = self.book.sheet_by_name(self.book.sheet_names()[0])
        #self.sheet = self.book.sheet_by_name('图纸')
        #self.rows_no = self.sheet.nrows
        #self.cols_no = self.sheet.ncols
'''
    def write_new_drawing_info(self, drawing_info):
        temp_book = copy(self.book)  # 利用xlutils.copy下的copy函数复制
        temp_sheet = temp_book.get_sheet(0)
        for i in range(self.rows_no-1):
            if self.sheet.cell(i, 3).value == drawing_info['drawing_no']:
                temp_sheet.write(i, 5, 'OK')
        temp_book.save(self.address)  # 保存文件

    def read_drawing_info():
        """
        Slot documentation goes here.
        """
        # TODO: not implemented yet
        raise NotImplementedError

    def drawing_content_update(self):
        """
        Slot documentation goes here.
        """
        # TODO: not implemented yet
        raise NotImplementedError
'''
'''
    def read_group_info(self):
        workbook = xlrd.open_workbook(workbook_address)
        sheet1_name = workbook.sheet_names()[0]
    # 根据sheet索引或者名称获取sheet内容
    #sheet1 = workbook.sheet_by_index(0)  # 根据sheet索引获取sheet内容
        sheet1 = workbook.sheet_by_name(sheet1_name)# 根据sheet名称获取sheet内容
    #print(sheet1.name,sheet1.nrows,sheet1.ncols)# sheet的名称，行数，列数

    # 获取整行和整列的值（数组）
        rows = sheet1.row_values(1) # 获取第四行内容
        cols = sheet1.col_values(1) # 获取第三列内容
        print(rows)
        print(cols)

    # 获取单元格内容
        print(sheet1.cell(1,1).value)
    #print sheet2.cell_value(1,0).encode('utf-8')
    #print sheet2.row(1)[0].value.encode('utf-8')

    # 获取单元格内容的数据类型
        print(sheet1.cell(1,1).ctype)
'''

def main(args):
    return 0

if __name__ == '__main__':
    import sys, os


    folder_path = os.path.abspath(os.path.dirname(os.getcwd()))
    database_address = str(folder_path + "/data/drawings.xlsx").replace("\\", "/")
    test_drawing_info = {'drawing_no':'C0101'}

    test_run = Handle_drawing_info(database_address)
    #test_run.write_new_drawing_info(test_drawing_info)
    print(test_run.rows_no, test_run.cols_no)

    sys.exit(main(sys.argv))
