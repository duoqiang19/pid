# pid - handle_new_review_proccess
# 2020/1/31 - 22:59
# Copyright 2020 Wang D.Q <Wang D.Q@PyCharm>
# !/venv python
# -*- coding: utf-8 -*-

# import xlrd
# from xlutils.copy import copy  # 导入copy模块
import openpyxl

class Handle_new_review_proccess():

    def __init__(self, address):
        self.address = address
        self.book = openpyxl.load_workbook(self.address)
        self.sheet = self.book['出图详细']
        self.rows_no = self.sheet.max_row
        self.cols_no = self.sheet.max_column

    def write_new_review_proccess(self, drawing_info):
        temp_book = copy(self.book)  # 利用xlutils.copy下的copy函数复制
        temp_sheet = temp_book.get_sheet(0)
        for i in range(self.rows_no-1):
            if self.sheet.cell(i, 3).value == drawing_info['drawing_no']:
                temp_sheet.write(i, 5, 'OK')
        temp_book.save(self.address)  # 保存文件

    def read_new_review_proccess():
        """
        Slot documentation goes here.
        """
        # TODO: not implemented yet
        raise NotImplementedError

    def drawing_content_update(self):
        """
        Slot documentation goes here.
        """
        # TODO: not implemented yet
        raise NotImplementedError

if __name__ == '__main__':
    import sys, os


    folder_path = os.path.abspath(os.path.dirname(os.getcwd()))
    database_address = str(folder_path + "/data/drawings.xlsx").replace("\\", "/")
    test_drawing_info = {'drawing_no':'C0101'}

    test_run = Handle_new_review_proccess(database_address)
    #test_run.write_new_drawing_info(test_drawing_info)
    print(test_run.rows_no, test_run.cols_no)
    for row in test_run.sheet.rows:
        for cell in row:
            print(cell.value, end = ' ')


    sys.exit(main(sys.argv))
